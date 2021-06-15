#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
import datetime
from os.path import isfile
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, colors, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import *

class SimpleNewExcel:
    def newexcel(self):
        wb = Workbook()
        # 获取活动工作表
        ws = wb.active
        # 数据可以直接分配到单元格
        ws['A1'] = 42
        # 行也可以追加
        ws.append([1, 2, 3])
        # Python类型将自动转换
        ws['A2'] = datetime.datetime.now()
        # 保存文件
        wb.save("sample.xlsx")

class ExcelToDict:
    """
    将Excel文件对象转成Python字典对象
    """

    def __init__(self, file_dir, title_row=0):
        # 工作簿文件的路径
        self.file_dir = file_dir
        # 标题行位于第几行
        self.title_row = int(title_row)
        self.data_dict = {}
        self.work_book = None

    def open_object(self):
        """打开工作簿对象"""
        valid = isfile(self.file_dir)
        # file_dir指向的文件是否不存在
        if not valid:
            raise Exception('文件路径 {0} 不存在'.format(self.file_dir))
        self.work_book = load_workbook(filename=self.file_dir)

    def read_excel(self):
        """读取工作簿数据"""
        if not self.work_book:
            raise Exception('需要先调用 open_object() 方法以打开工作簿对象')
        for sheet_name in self.work_book.sheetnames:
            # 每个工作表的字典
            data_dict_sheet = {'title_row': [], 'value_row': {}}
            # 获取工作表对象
            ws = self.work_book[sheet_name]
            # 预先创建工作表中每一行的字典
            for i in range(ws.max_row - 1 - self.title_row):
                data_dict_sheet['value_row'][i] = {}
            # 遍历所有列
            columns = tuple(ws.columns)
            for column in columns:
                # 每一列的标题
                title = column[self.title_row].value
                # 记录每列的标题
                data_dict_sheet['title_row'].append(title)
                row_num = 0
                # 遍历每一列中的所有值
                for col in column:
                    # 忽略每一列的标题行
                    if column.index(col) <= self.title_row:
                        continue
                    data_dict_sheet['value_row'][row_num][title] = col.value
                    row_num += 1
            # 记录每个工作表的数据字段
            self.data_dict[sheet_name] = data_dict_sheet

    def check(self, check_item=None, sheet_name=None, sheet_index=0):
        """
        在所选工作表中校验是否包含业务需要的所有标题名称
        :param check_item: 所选工作表中需要校验的标题列表
        :param sheet_name: 以名称形式选择工作表（优先）
        :param sheet_index: 以下标形式选择工作表
        :return: {'result': True, 'exception': None}
        """
        if not self.data_dict:
            return {'result': False, 'exception': '需要先调用 read_excel() 方法以读取工作簿数据'}
        if check_item is None:
            check_item = []
        if sheet_name:
            if sheet_name not in self.data_dict:
                return {'result': False, 'exception': '不存在名为 {0} 的工作表'.format(sheet_name)}
            # 直接获得对应的工作表数据
            data_sheet = self.data_dict[sheet_name]
        else:
            # 通过下标获取对应的工作表名称
            data_dict_keys = tuple(self.data_dict.keys())
            if len(data_dict_keys) <= int(sheet_index):
                return {'result': False, 'exception': '不存在下标为 {0} 的工作表'.format(sheet_index)}
            _sheet_name = data_dict_keys[int(sheet_index)]
            # 间接获得对应的工作表数据
            data_sheet = self.data_dict[_sheet_name]
        # 判断工作表中是否包含业务需要的所有标题
        if not set(check_item).issubset(set(data_sheet['title_row'])):
            return {'result': False, 'exception': '工作表中未包含业务需要的 {0} 标题'.format(check_item)}
        return {'result': True, 'exception': None}


class Demo:
    def newexcel(self):
        # 创建工作表
        wb = Workbook()
        ws = wb.active

        # 打开已有工作表
        wb2 = load_workbook('文件名称.xlsx')

        # 创建工作表页面sheel在末尾插入（默认）
        ws1 = wb.create_sheet("Mysheet")
        # 插入第一个位置
        ws2 = wb.create_sheet("Mysheet", 0)
        # 倒数第二个位置插入
        ws3 = wb.create_sheet("Mysheet", -1)

        # 更改工作表名称
        ws.title = "New Title"

        # 获得工作表对象
        ws4 = wb["New Title"]
        ws5 = wb.get_sheet_by_name("New Title")

        # 查看所有工作表
        print(wb.sheetnames)  # ['Sheet2', 'New Title', 'Sheet1']
        for sheet in wb:
            print(sheet.title)

        # 复制工作表
        source = wb.active
        target = wb.copy_worksheet(source)

        # 单元格操作，键值进行访问
        c = ws['A4']
        ws['A4'] = 4

        #  行 和 列 定位要访问的单元格
        d = ws.cell(row=4, column=2, value=10)

        # 多单元格，键值操作
        cell_range = ws['A1':'C2']
        colC = ws['C']
        col_range = ws['C:D']
        row10 = ws[10]
        row_range = ws[5:10]

        # Worksheet.iter_rows() 或 Worksheet.iter_cols() 方法获取行、列。循环复杂获得
        for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
            for cell in row:
                print(cell)

        # 遍历所有行列
        ws = wb.active
        ws['C9'] = 'hello world'
        tuple(ws.rows)
        tuple(ws.columns)

        # 仅值，通过 Worksheet.iter_rows() 并 Worksheet.iter_cols() 可以获取 values_only 参数，只返回单元格的值：
        for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
            print(row)
        # (None, None, None)
        # (None, None, None)

        # 工作表最行列数
        print(ws.max_row)  # 4
        print(ws.max_column)  # 15

        # 数据存储
        ws3.cell(column=2, row=2, value="asd")
        c.value = 'hello, world'
        print(c.value)  # 'hello, world'
        d.value = 3.14
        print(d.value)  # 3.14
        wb.save('balances.xlsx')

        # 根据列的数字返回字母
        print(get_column_letter(2))  # B
        # 根据字母返回列的数字
        print(column_index_from_string('D'))  # 4

        # 将文件保存到流中，例如在使用Web应用程序（Pyramid、Flask、Django）时，只需使用 NamedTemporaryFile() 方法即可：
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        # 使用数字格式日期
        ws['A1'] = datetime.datetime(2010, 7, 21)
        ws['A1'].number_format  # 'yyyy-mm-dd h:mm:ss'

        # 添加一个简单的公式
        ws["A1"] = "=SUM(1, 1)"
        wb.save("formula.xlsx")

        # 创建图像
        img = Image('logo.png')
        # 添加到工作表并锚定在单元格旁边
        ws.add_image(img, 'A1')

        # 删除表
        wb.remove(ws)

        # 设置单元格风格
        # 字体
        bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
        ws['A1'].font = bold_itatic_24_font

        # 设置B1中的数据垂直居中和水平居中
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

        # 第2行行高
        ws.row_dimensions[2].height = 40
        # C列列宽
        ws.column_dimensions['C'].width = 30

        # 合并单元格， 往左上角写入数据即可
        ws.merge_cells('B1:G1')  # 合并一行中的几个单元格
        ws.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格
        # 合并后只可以往左上角写入数据，也就是区间中: 左边的坐标。
        # 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
        # 以下是拆分单元格的代码。拆分后，值回到A1位置
        ws.unmerge_cells('A1:C3')


if __name__ == '__main__':
    excel_to_dict = ExcelToDict('C:/Users/hekaiyou/Desktop/新建 Microsoft Excel 工作表.xlsx')
    excel_to_dict.open_object()
    print('工作簿对象：', excel_to_dict.work_book)
    excel_to_dict.read_excel()
    print('工作簿数据：', excel_to_dict.data_dict)
    print('工作簿校验（异常演示）：', excel_to_dict.check(['标题四']))
    print('工作簿校验（正常演示）：', excel_to_dict.check(['标题一', '标题二']))
